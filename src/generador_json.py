import json

from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

ZONA_HORARIA_ECUADOR = ZoneInfo(
    "America/Guayaquil"
)


# ============================================================
# CONVERTIR VALORES DE EXCEL A JSON
# ============================================================


def convertir_valor_json(valor):

    # --------------------------------------------------------
    # FECHA + HORA
    # --------------------------------------------------------

    if isinstance(
        valor,
        datetime
    ):

        return valor.isoformat()

    # --------------------------------------------------------
    # FECHA
    # --------------------------------------------------------

    if isinstance(
        valor,
        date
    ):

        return valor.isoformat()

    # --------------------------------------------------------
    # RESTO DE VALORES
    # --------------------------------------------------------

    return valor


# ============================================================
# GENERAR JSON DESDE EXCEL
# ============================================================


def generar_json_desde_excel(
    archivo_excel,
    reporte_id,
    reporte_nombre,
    fecha_desde="2026-01-01",
):

    print("")
    print("==========================================")
    print(" GENERANDO JSON")
    print("==========================================")

    archivo_excel = Path(
        archivo_excel
    )

    # --------------------------------------------------------
    # 1. VALIDAR EXCEL
    # --------------------------------------------------------

    if not archivo_excel.exists():

        raise FileNotFoundError(
            f"No existe el Excel: "
            f"{archivo_excel}"
        )

    if archivo_excel.stat().st_size <= 0:

        raise RuntimeError(
            "El Excel está vacío."
        )

    print("Reporte:")
    print(
        f"ID {reporte_id} - "
        f"{reporte_nombre}"
    )

    print("Archivo Excel:")
    print(
        archivo_excel
    )

    # --------------------------------------------------------
    # 2. ABRIR EXCEL
    #
    # data_only=True permite trabajar con los valores
    # calculados en lugar de las fórmulas cuando existan.
    # --------------------------------------------------------

    workbook = load_workbook(
        filename=archivo_excel,
        read_only=True,
        data_only=True
    )

    # --------------------------------------------------------
    # 3. UTILIZAR HOJA DATA
    # --------------------------------------------------------

    if "Data" in workbook.sheetnames:

        worksheet = workbook[
            "Data"
        ]

    else:

        worksheet = workbook.active

    print("Hoja utilizada:")
    print(
        worksheet.title
    )

    # --------------------------------------------------------
    # 4. LEER ENCABEZADOS
    # --------------------------------------------------------

    filas = worksheet.iter_rows(
        values_only=True
    )

    try:

        encabezados_originales = next(
            filas
        )

    except StopIteration:

        workbook.close()

        raise RuntimeError(
            "El Excel no contiene filas."
        )

    columnas = []

    for indice, encabezado in enumerate(
        encabezados_originales,
        start=1
    ):

        if encabezado is None:

            nombre_columna = (
                f"Columna_{indice}"
            )

        else:

            nombre_columna = str(
                encabezado
            ).strip()

        # Evitar claves JSON duplicadas
        nombre_base = nombre_columna
        contador = 2

        while nombre_columna in columnas:

            nombre_columna = (
                f"{nombre_base}_{contador}"
            )

            contador += 1

        columnas.append(
            nombre_columna
        )

    # --------------------------------------------------------
    # 5. CONSTRUIR REGISTROS
    # --------------------------------------------------------

    registros = []

    for fila in filas:

        # Ignorar filas totalmente vacías
        if all(
            valor is None
            for valor in fila
        ):

            continue

        registro = {}

        for indice, columna in enumerate(
            columnas
        ):

            if indice < len(fila):

                valor = fila[indice]

            else:

                valor = None

            registro[columna] = (
                convertir_valor_json(
                    valor
                )
            )

        registros.append(
            registro
        )

    workbook.close()

    # --------------------------------------------------------
    # 6. FECHA DE GENERACIÓN
    # --------------------------------------------------------

    momento_ecuador = datetime.now(
        ZONA_HORARIA_ECUADOR
    )

    fecha_hasta = (
        momento_ecuador
        .strftime("%Y-%m-%d")
    )

    fecha_generacion = (
        momento_ecuador.isoformat(
            timespec="seconds"
        )
    )

    # --------------------------------------------------------
    # 7. ESTRUCTURA JSON
    # --------------------------------------------------------

    contenido_json = {

        "metadata": {

            "reporte_id": (
                reporte_id
            ),

            "reporte_nombre": (
                reporte_nombre
            ),

            "fecha_desde": (
                fecha_desde
            ),

            "fecha_hasta": (
                fecha_hasta
            ),

            "fecha_generacion": (
                fecha_generacion
            ),

            "zona_horaria": (
                "America/Guayaquil"
            ),

            "archivo_origen": (
                archivo_excel.name
            ),

            "hoja_origen": (
                worksheet.title
            ),

            "total_registros": (
                len(registros)
            ),

            "total_columnas": (
                len(columnas)
            ),
        },

        "columnas": columnas,

        "registros": registros,
    }

    # --------------------------------------------------------
    # 8. NOMBRE DEL JSON
    #
    # Conservamos exactamente el mismo nombre base
    # que el Excel.
    # --------------------------------------------------------

    archivo_json = (
        archivo_excel.with_suffix(
            ".json"
        )
    )

    # --------------------------------------------------------
    # 9. GUARDAR JSON
    # --------------------------------------------------------

    with open(
        archivo_json,
        "w",
        encoding="utf-8"
    ) as archivo:

        json.dump(
            contenido_json,
            archivo,
            ensure_ascii=False,
            indent=2
        )

    # --------------------------------------------------------
    # 10. VALIDAR JSON
    # --------------------------------------------------------

    if not archivo_json.exists():

        raise RuntimeError(
            "El JSON no fue generado."
        )

    if archivo_json.stat().st_size <= 0:

        raise RuntimeError(
            "El JSON generado está vacío."
        )

    print("")
    print(
        "✅ JSON GENERADO CORRECTAMENTE"
    )

    print("Archivo:")
    print(
        archivo_json
    )

    print("Registros:")
    print(
        len(registros)
    )

    print("Columnas:")
    print(
        len(columnas)
    )

    print("Tamaño:")
    print(
        archivo_json.stat().st_size,
        "bytes"
    )

    return archivo_json